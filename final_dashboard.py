import sys
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import QFileInfo
from openpyxl import load_workbook
from PyQt5.QtWidgets import QHeaderView, QTableWidget, QLabel, QMessageBox, QComboBox, QCheckBox, QHBoxLayout, QTableWidgetItem, QFileDialog, QAction, QStackedLayout, QMainWindow, QWidget
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtCore import QFile

firstTabDone = False

def import_file():
    global firstTabDone
    if firstTabDone:
        global_file.table_obj.add_tabs()
       
    firstTabDone = True
    file_dialog = QFileDialog()
    file_dialog.setFileMode(QFileDialog.ExistingFile)
    file_dialog.setNameFilter("Excel Files (*.xlsx)")
    
    if file_dialog.exec_():
        file_path = file_dialog.selectedFiles()[0]
        global_file.table_obj.tab_name_change(QFileInfo(file_path).fileName())
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            rows = list(filter(None, rows))[1:]
            num_rows = len(rows)
            num_cols = sheet.max_column
            
            global_file.table_obj.original_table.setRowCount(num_rows)
            global_file.table_obj.original_table.setColumnCount(num_cols)
            
            for row_index, row_data in enumerate(rows):
                for col_index, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value))
                    global_file.table_obj.original_table.setItem(row_index, col_index, item)
                    
            headers = []
            for col in range(1, sheet.max_column+1):
                header = sheet.cell(row=1, column=col).value
                headers.append(header)
                
            global_file.table_obj.original_table.setHorizontalHeaderLabels(headers)
            horizontal_header = global_file.table_obj.original_table.horizontalHeader()
            horizontal_header.setSectionResizeMode(QHeaderView.Stretch)
            horizontal_header.setStretchLastSection(True)
            
            # FILTERS updated in the left layout
            global_file.table_obj.checkbox_temp_list = []
            global_file.table_obj.combobox_temp_list = []
            for col in range(sheet.max_column+1):
                isInt = False
                global_file.table_obj.filterHbox = QHBoxLayout()
                header_item = global_file.table_obj.original_table.horizontalHeaderItem(col)
                
                if header_item is not None and (header_item in global_file.table_obj.cbox_filter_names):
                    if not header_item.text() == "Txn Start time (ns)" and not header_item.text() == "Txn End time (ns)":
                        global_file.table_obj.checkbox_temp = QCheckBox(header_item.text())
                        global_file.table_obj.filterHbox.addWidget(global_file.table_obj.checkbox_temp)
                        global_file.table_obj.checkbox_temp_list.append(global_file.table_obj.checkbox_temp)
                        global_file.table_obj.options = QComboBox()
                        global_file.table_obj.options.addItem('--SELECT--')
                        global_file.table_obj.filterHbox.addWidget(global_file.table_obj.options)
                        global_file.table_obj.combobox_temp_list.append(global_file.table_obj.options)
                        global_file.table_obj.options.setEnabled(False)
                        global_file.table_obj.leftLayout.addLayout(global_file.table_obj.filterHbox)
                        global_file.table_obj.filter_temp_list = []
                        
                        for row in range(num_rows):
                            item = QTableWidgetItem(str(global_file.table_obj.original_table.item(row, col).text()))
                            global_file.table_obj.original_table.setItem(row, col, item);
                            if not str(global_file.table_obj.original_table.item(row, col).text()) in global_file.table_obj.filter_temp_list:
                                global_file.table_obj.filter_temp_list.append(str(global_file.table_obj.original_table.item(row, col).text()))
                                if(str(global_file.table_obj.original_table.item(row, col)).isdigit()):
                                    isInt = True
                        if isinstance(global_file.table_obj.options, QComboBox):
                            if isInt:
                                sorted_list = sorted(global_file.table_obj.filter_temp_list, key=lambda x: int(x))
                                for i in range(len(sorted_list)):
                                    global_file.table_obj.options.addItem(str(sorted_list[i]))
                            else:
                                for i in range(len(global_file.table_obj.filter_temp_list)):
                                    global_file.table_obj.options.addItem(str(global_file.table_obj.filter_temp_list[i]))
                
        except Exception as e:
            print(f"Error: {str(e)}")
    else:
        print("No file selected.")

def compare_action_triggered():
    stacked_layout.setCurrentWidget(global_file.compare_obj)

def import_action_triggered():
    import_file()

def exit_action_triggered():
    confirm_dialog = QMessageBox()
    confirm_dialog.setIcon(QMessageBox.Warning)
    confirm_dialog.setText("Are you sure you want to exit the application?")
    confirm_dialog.setWindowTitle("Confirmation")
    confirm_dialog.setStandardButtons(QMessageBox.Yes | QMessageBox.No)

    # If the user confirms, remove the tab
    if confirm_dialog.exec_() == QMessageBox.Yes:
        app.quit()

def view_action_triggered():
    stacked_layout.setCurrentWidget(global_file.table_obj)

def home_action_triggered():
    try:
        stacked_layout.setCurrentWidget(global_file.dashboard_obj)
        global_file.dashboard_obj.clear_layout(global_file.dashboard_obj.left_stack_child)
        if not global_file.dashboard_obj.tabs_bottom.count() == 0:
            table = global_file.dashboard_obj.tabs_bottom.widget(global_file.dashboard_obj.tab_index).findChild(QTableWidget)
            if not global_file.dashboard_obj.left_stack_child.count() == 0:
                global_file.dashboard_obj.clear_layout(global_file.dashboard_obj.left_stack_child)
            else:
                global_file.dashboard_obj.left_stack.addLayout(global_file.dashboard_obj.left_stack_child)
                global_file.dashboard_obj.left_stack_child.addWidget(QLabel(global_file.dashboard_obj.statistic.read_write_count(table)))
                global_file.dashboard_obj.left_stack_child.addWidget(QLabel('Average data bytes transferred: {:.2f} bytes'.format(global_file.dashboard_obj.statistic.data_bytes_transfer_avg(table))))
                global_file.dashboard_obj.left_stack_child.addWidget(QLabel('Average transaction time: {:.2f} ns'.format(global_file.dashboard_obj.statistic.txn_time_diff_avg(table))))
                global_file.dashboard_obj.left_stack_child.addWidget(QLabel(global_file.dashboard_obj.statistic.total_fixed_incr(table)))
                bd_label = QLabel(global_file.dashboard_obj.statistic.bandwidthInfo(table))
                font = QFont()
                font.setWeight(QFont.Bold)
                bd_label.setFont(font)
                global_file.dashboard_obj.left_stack_child.addWidget(bd_label)
    except TypeError:
        #QMessageBox.information(None, "FATAL ERROR", "some of the stats are not visible due to incorrect naming or missing columns!") 
        pass
    
def help_action_triggered():
    help_info = """
FILE:
Import File: You can import only .xlsx file, but multiple number of times.
Exit Application: It will exit the application.

TABLE:
View: Any imported table and its filters will be seen in this section. You can apply filters to the table and "Export to dashboard" for further analysis. If you want to add more filters then you have to remove the existing filters and add new filters again! Once a table is exported to the dashboard, you can see it by navigating to Dashboard -> Home.

DASHBOARD:
Home: You can see all the imported tables here and you can make plot of the selected table based on the options given to you on the left side of the page. To import/filter more data navigate back to Table -> View.

{ERROR}:
If the table does not have expected column names, then you can face issues with graphical and statistical analysis!!
"""
    QMessageBox().information(None, 'INFORMATION', help_info)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    import global_file
    stacked_layout = QStackedLayout()
    #style_file = QFile("style.qss")
    #style_file.open(QFile.ReadOnly | QFile.Text)
    #style_sheet = style_file.readAll()
    
    # Set the style sheet
    #app.setStyleSheet(str(style_sheet, encoding='utf-8'))
    
    mainLayout = QMainWindow()
    menu_bar = mainLayout.menuBar()
    file_menu = menu_bar.addMenu('File')
    table_menu = menu_bar.addMenu('Table')
    dashboard_menu = menu_bar.addMenu('Dashboard')
    help_menu = menu_bar.addMenu('Help')
    mainLayout.setMenuBar(menu_bar)
    
    import_action = QAction(QIcon(r"C:\Users\z004suej\Desktop\app\images\file2.png"), "Import File")
    exit_action = QAction(QIcon(r"C:\Users\z004suej\Desktop\app\images\exit.png"), "Exit Application")
    view_action = QAction(QIcon(r"C:\Users\z004suej\Desktop\app\images\table.png"), "View")
    compare_action = QAction(QIcon(r"C:\Users\z004suej\Desktop\app\images\compare.png"), "Compare Data")
    home_action = QAction(QIcon(r"C:\Users\z004suej\Desktop\app\images\dashboard.png"), "Home")
    help_action = QAction(QIcon(r"C:\Users\z004suej\Desktop\app\images\help.png"), "Help")
    
    file_menu.addAction(import_action)
    file_menu.addAction(exit_action)
    table_menu.addAction(view_action)
    dashboard_menu.addAction(home_action)
    dashboard_menu.addAction(compare_action)
    help_menu.addAction(help_action)
    
    import_action.triggered.connect(import_action_triggered)
    exit_action.triggered.connect(exit_action_triggered)
    view_action.triggered.connect(view_action_triggered)
    compare_action.triggered.connect(compare_action_triggered)
    home_action.triggered.connect(home_action_triggered)
    help_action.triggered.connect(help_action_triggered)
    
    central_widget = QWidget()
    mainLayout.setCentralWidget(central_widget)
    central_widget.setLayout(stacked_layout)
    stacked_layout.addWidget(global_file.table_obj)
    stacked_layout.addWidget(global_file.dashboard_obj)
    stacked_layout.addWidget(global_file.compare_obj)
    
    stacked_layout.setCurrentWidget(global_file.table_obj)
    mainLayout.setWindowTitle('AXI-BUS Communication Data Analysis & Prediction')
    mainLayout.showMaximized()
    app.exec_()
    sys.exit()